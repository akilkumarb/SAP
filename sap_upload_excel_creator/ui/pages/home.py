from pathlib import Path
import pandas as pd
import streamlit as st
import datetime
import time
import pickle

# from sap_upload_excel_creator.core.validate import check_mandatory_columns_available
from ...core.validate import check_mandatory_columns_available
from ...core.mappers import load_relevant_sheet_names_mapping
from ...core.functions import generate, dataframe_export_to_excel

from ...paths import CRED_DIR

import hashlib

def hash_password(password):
    # Use sha256 to hash the password
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    return password_hash


# MAPPING_SEC_PASSWORD_IS_CORRECT = False
def login_mapping_section():
    password_hash_file_path = CRED_DIR / 'mapping_page_login_password_hash'
    with open(password_hash_file_path, "rb") as f:
        password_hash = pickle.load(f)

    st.caption("Please enter the password to view or change the mapping")
    password = st.text_input("Password", type='password')
    submit = st.button("Submit")

    if submit and hash_password(password) == password_hash:
        st.success("Login successful!")
        st.session_state['MAPPING_SEC_PASSWORD_IS_CORRECT'] = True
    elif submit and hash_password(password) != password_hash:
        st.error("Incorrect password.")


def render_home():
    if 'MAPPING_SEC_PASSWORD_IS_CORRECT' not in st.session_state: 
        st.session_state['MAPPING_SEC_PASSWORD_IS_CORRECT'] = False
    # global MAPPING_SEC_PASSWORD_IS_CORRECT
    st.title('SAP Upload Data Creator')
    st.subheader('An Ezetap Tool to aggregate excel files from sources to create output excel files for SAP')
    main_tab, mapping_tab, settings_tab = st.tabs(["Generate", "Mapping", "Settings"])
    with main_tab:
        render_form_tab()
    with mapping_tab:
        st.subheader("Upload and change existing Data Mappings")
        login_mapping_section()
        if st.session_state['MAPPING_SEC_PASSWORD_IS_CORRECT'] is True:
            render_mapping_tab()

    with settings_tab:
        render_settings_tab()
        

SUBMISSION_COMPLETED = False
GLOBAL_DFS = dict()
DATE_FOR_FILTER = None
FINAL_DF = None
DOWNLOAD_READY = False

def render_form_tab():
    st.info("Always make sure that you refresh the web page before the next set of file uploads. ")  # MUST TELL if file is closed and added another file one load button is pressed.
    global SUBMISSION_COMPLETED, GLOBAL_DFS, DATE_FOR_FILTER, FINAL_DF, DOWNLOAD_READY
    # SUBMISSION_COMPLETED = False
    # GLOBAL_DFS = dict()
    # DATE_FOR_FILTER = None
    # FINAL_DF = None
    # DOWNLOAD_READY = False
    with st.form("Upload Section"):
        st.write("Please fill in the following input fields:")

        DATE_FOR_FILTER = st.date_input(
            "Select the Date for filtering data",
            datetime.datetime.now().date())
        st.write('Selected Date is:', DATE_FOR_FILTER.strftime("%d/%m/%Y"))

        col1, col2, col3 = st.columns(3)
        with col1:
            kotak_uploaded_file = st.file_uploader("Kotak Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
            axis_uploaded_file = st.file_uploader("Axis Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
            ybl_uploaded_file = st.file_uploader("YBL Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
        
        with col2:
            sbi_uploaded_file = st.file_uploader("SBI Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
            hdfc_uploaded_file = st.file_uploader("HDFC Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
            hdfc_aggr_uploaded_file = st.file_uploader("HDFC Aggr Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
        
        with col3:
            icici_uploaded_file = st.file_uploader("ICICI Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
            idfc_uploaded_file = st.file_uploader("IDFC Master Excel: ", type=['xlsx', 'xlsb'], accept_multiple_files=False)
            complaints_uploaded_file = st.file_uploader("Complaints Master Excel", type=['xlsx', 'xlsb'], accept_multiple_files=False)


        


    #    slider_val = st.slider("Form slider")
    #    checkbox_val = st.checkbox("Form checkbox")

        # Every form must have a submit button.
        submitted = st.form_submit_button("Load and Validate")
        uploaded_files = {
            "KOTAK": kotak_uploaded_file, 
            "AXIS": axis_uploaded_file, 
            "YBL": ybl_uploaded_file,
            "SBI": sbi_uploaded_file, 
            "HDFC": hdfc_uploaded_file, 
            "HDFC AGGR": hdfc_aggr_uploaded_file,
            "ICICI": icici_uploaded_file, 
            "IDFC": idfc_uploaded_file, 
            "COMPLAINTS": complaints_uploaded_file
        }
        
        if submitted:
            
            relevant_sheet_names = load_relevant_sheet_names_mapping()
            empty_fields = [k for k, v in  uploaded_files.items() if v is None]
            if "COMPLAINTS" in empty_fields:
                st.error(f"The following excel data is not uploaded: {empty_fields}. ")
            elif empty_fields:
                st.warning(f"WARNING! Excel data for the following banks are not uploaded: {empty_fields}")
            
            if list(set(uploaded_files.keys()) - set(empty_fields)):
                with st.spinner('Loading & Validating...'):
                    
                    # excluding if dataframe is not avalable (value is None)
                    selected_files = {key: value for key, value in uploaded_files.items() if value is not None}

                    # st.write(selected_files)   # # FOR DEBUG PURPOSE IN UI

                    # loading all dataframes
                    # dfs = dict()
                    MANDATORY_DATA_SHEET_NOT_FOUND = list()
                    currently_selected_ids = []   # to avoid previously loaded not currenlty avaiable file_id datas THIS IS A BUG FIX
                    for file_id, filename in selected_files.items():
                        print(f"{'Reading file:':<20} {file_id:>20}")
                        sheets = pd.read_excel(filename, sheet_name=None)
                        if relevant_sheet_names[file_id] not in sheets.keys():
                            MANDATORY_DATA_SHEET_NOT_FOUND.append((file_id, relevant_sheet_names[file_id]))
                        else:
                            GLOBAL_DFS[file_id] = sheets[relevant_sheet_names[file_id]]
                            print(f"Done".rjust(40, " "))
                            print(40*"-")
                            currently_selected_ids.append(file_id)
                    
                    # THIS IS TO FIX THE BUG
                    for previously_available_file_id in GLOBAL_DFS.keys():  # previously available in the previous sessioon
                        if previously_available_file_id not in currently_selected_ids:
                            GLOBAL_DFS[previously_available_file_id] = None  # set value to None OR remove entire item key and value
                            # if del is used, the dictionary item deleted in iteration error will be raised. so set to None for now
                    
                    if MANDATORY_DATA_SHEET_NOT_FOUND:
                        er_msg = ", ".join([f"{i[0]} MASTER's sheet named '{i[1]}'" for i in MANDATORY_DATA_SHEET_NOT_FOUND]) 
                        message = f"The mandatory sheets are not found: {er_msg}"
                        print(message)
                        data = None
                        is_valid = False

                    else:
                        # st.write("after loading\n", GLOBAL_DFS)    # FOR DEBUG PURPOSE IN UI

                        # checking if all mandatory columns available
                        is_valid, message, data = check_mandatory_columns_available(GLOBAL_DFS)
            
                if is_valid:
                    # st.success(message + str(data))
                    SUBMISSION_COMPLETED = True
                else:
                    st.error(message)
                    if data is not None:
                        for file_id in data.keys():
                            st.error(f"{file_id}: {data[file_id]}")

    if SUBMISSION_COMPLETED is True:
        print("Successfully Submitted and Loded into memmory")
        loaded_final_df = None
        if st.button("Generate"):
            FINAL_DF = generate(GLOBAL_DFS, date_for_data_filter=DATE_FOR_FILTER)
            loaded_final_df = FINAL_DF.copy()  # copying to different variable to make FINAL_DF None so avoid reload issues
            FINAL_DF = None

        if loaded_final_df is not None:
            st.dataframe(loaded_final_df)
            DOWNLOAD_READY = True
            st.download_button(label="Download the aggregated excel sheet", file_name='SAP_READY_MASTER.xlsx', mime=None, data=get_output_excel_data(loaded_final_df), disabled=not DOWNLOAD_READY)
            SUBMISSION_COMPLETED = False




def get_output_excel_data(loaded_final_df):
    master_output_file = Path("MASTER.xlsx")
    if master_output_file.is_file():
        print("Deleting previous file")
        master_output_file.unlink()

    dataframe_export_to_excel(loaded_final_df, master_output_file, sheet_name="Data Aggregated", engine="xlsxwriter")
    # pd.DataFrame([pd.Series(data.JSON_DICT)]).T.to_csv(str(config.TEMPORARY_CSV_PATH))
    return open(str(master_output_file), 'rb')


MAPPING_LOGIN_SUBMIT = False

def render_mapping_tab():
    

    # st.caption("Please enter the password to view or change the mapping")
    # password = st.text_input("Password", type='password')
    # submit = st.button("Submit")

    # if submit and password == "secret":
    #     st.success("Login successful!")
        # download_ready = False
        # col1, col2, col3 = st.columns(3)
    if st.button("View the Existing Mapping data"):
        # st.dataframe(loaded_final_df)
        mapping_excel_path, msg, sheets = create_xl_mapping_file()
        if sheets:
            # (main_tab, mapping_tab, settings_tab)= st.tabs(list(sheets.keys()))
            for sht_name, sht in sheets.items():
                with st.expander(f"{sht_name}", expanded=False):
                    st.write(sht)
        if mapping_excel_path is not None:
            if mapping_excel_path.is_file():
                download_ready = True
                # st.success(msg)
                st.download_button(
                    label="Download the Existing Mappings", 
                    file_name='MAPPING.xlsx', 
                    mime=None, 
                    data=open(str(mapping_excel_path), 'rb'), 
                    disabled=not download_ready
                )
            else:
                st.error("Some error! The converted file is not found")
        else:
            st.error(msg)
        # download_ready = True
    # upload mappging excel file section
    with st.form("Load New Mapping"):
        uploaded_mapping_xl_file = st.file_uploader("Upload Mapping File", type=['xlsx',], accept_multiple_files=False)
        submitted = st.form_submit_button("Submit the Mapping File")
        if submitted:
            if uploaded_mapping_xl_file is None:
                st.warning("Please upload mapping excel and try clicking")
            else:
                with st.spinner('Loading..'):
                    successful, message = load_uploaded_excel_2_json_mappings(uploaded_mapping_xl_file)
                    if successful:
                        st.success(message)
                    else:
                        st.error(message)
    

    


import json
from pathlib import Path
from ...paths import MAPPINGS_DIR, TEMP_DIR
def load_uploaded_excel_2_json_mappings(uploaded_mapping_xl_file):
    json_excel_sheet_mapper_path = MAPPINGS_DIR / "json_excel_sheet_mapper.json"
    json_excel_sheet_mapping =json.loads(json_excel_sheet_mapper_path.read_bytes())
    try:
        # sheets = pd.read_excel(uploaded_mapping_xl_file, sheet_name=None, engine="openpyxl")
        # sheet_names = list(sheets.keys())
        # del sheets
        # st.write(sheet_names)

        mappings_excel_filepath = uploaded_mapping_xl_file

        # LOAD NEW MAPPING (accepts excelsheet with defined sheet names and their defined columns)
        out_jsons = dict()   # so that individual writing won't happen and breaking then in another json writing


        key = "bp_codes_n_names"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0],
            engine="openpyxl"
        )
        return_json_data = df.T.to_dict()
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data


        key = "acquirer_n_device_model_to_item_no_n_item_desciption"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0, 1],
            engine="openpyxl"
        )
        df.reset_index(inplace=True)
        groups = df.groupby('device')
        level_1_unpacked_dict = {group_id: groups.get_group(group_id).drop(columns=['device']) for group_id in df['device'].unique()}
        return_json_data = {device: df.set_index("bank").to_dict(orient="index") for device, df in level_1_unpacked_dict.items()}
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data


        key = "deactivation_date_column_mapping"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0],
            engine="openpyxl"
        )
        return_json_data = df['Column Name'].to_dict()
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data



        key = "region_to_spl_codes"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0],
            engine="openpyxl"
        )
        return_json_data = df.T.to_dict()
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data



        key = "sheet_names"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0],
            engine="openpyxl"
        )
        return_json_data = df['Sheet Name'].to_dict()
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data



        key = "statenames_n_short_codes"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0],
            engine="openpyxl"
        )
        return_json_data = df["Short Code"].to_dict()
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data



        key = "status_columns"
        df = pd.read_excel(
            mappings_excel_filepath, 
            sheet_name=json_excel_sheet_mapping[key]['sheet_name'], 
            index_col=[0],
            engine="openpyxl"
        )
        return_json_data = df["Column Name"].to_dict()
        output_json_path = MAPPINGS_DIR / json_excel_sheet_mapping[key]['json_file']
        out_jsons[str(output_json_path)] = return_json_data


        # writing to json files 
        backup_dir = MAPPINGS_DIR / "backup"

        for path_str, json_data in out_jsons.items():
            # backing up the data
            path = Path(path_str)
            print("backing up", path.name)
            with open(backup_dir / path.name, 'w') as backup_file_object:
                backup_file_object.write(open(path, "r").read())
            
            print("Saving new json data to original mapping json -", path.name)
            with open(path, "w") as f:
                json.dump(json_data, f, sort_keys=False, indent=4)

        
        return True, "Success"
    except:
        return False, "Uploaded Excel file has something which is creating an issue to while processing"


def create_xl_mapping_file():
    try:
        json_excel_sheet_mapper_path = MAPPINGS_DIR / "json_excel_sheet_mapper.json"
        json_excel_sheet_mapping =json.loads(json_excel_sheet_mapper_path.read_bytes())

        sheets = dict()

        key = "bp_codes_n_names"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        sheet = pd.read_json(json_file_path).T
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet


        key = "acquirer_n_device_model_to_item_no_n_item_desciption"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        json_data =json.loads(json_file_path.read_bytes())
        first_columns = ['device', "bank"]
        ddfs = []
        for device in json_data.keys():
            df = pd.DataFrame(json_data[device]).T.reset_index().rename(columns={'index': "bank"})
            df['device'] = device
            df = df[first_columns + [col for col in df.columns if col not in first_columns]]
            ddfs.append(df)
        sheet = pd.concat(ddfs).reset_index(drop=True)
        sheet = sheet.set_index(first_columns)[['Item No.', 'Item Description']]
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet

        key = "deactivation_date_column_mapping"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        json_data =json.loads(json_file_path.read_bytes())
        sheet = pd.DataFrame([json_data]).T.rename(columns={0: "Column Name"})
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet


        key = "region_to_spl_codes"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        sheet = pd.read_json(json_file_path).T
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet


        key = "sheet_names"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        json_data =json.loads(json_file_path.read_bytes())
        sheet = pd.DataFrame([json_data]).T.rename(columns={0: "Sheet Name"})
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet


        key = "statenames_n_short_codes"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        json_data =json.loads(json_file_path.read_bytes())
        sheet = pd.DataFrame([json_data]).T.rename(columns={0: "Short Code"})
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet


        key = "status_columns"
        json_file_name = json_excel_sheet_mapping[key]["json_file"]
        json_file_path = MAPPINGS_DIR / json_file_name
        json_data =json.loads(json_file_path.read_bytes())
        sheet = pd.DataFrame([json_data]).T.rename(columns={0: "Column Name"})
        sheets[json_excel_sheet_mapping[key]["sheet_name"]] = sheet


        from openpyxl.utils import get_column_letter
        temporary_mapping_excel_file = TEMP_DIR / 'mappings.xlsx'
        # book = load_workbook('test.xlsx')
        writer = pd.ExcelWriter(temporary_mapping_excel_file, engine='openpyxl')
        for sheet_name, sheet_data in sheets.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name)
            max_col = sheet_data.shape[1]
            worksheet = writer.sheets[sheet_name]
            for col_number in range(1, max_col+3):
                worksheet.column_dimensions[get_column_letter(col_number)].width = 40

        writer.close()

        return temporary_mapping_excel_file, "Success", sheets
    except:
        sheets = dict()
        return None, "Unable to process the new excel file. Please make sure everything is properly added", sheets

def mapping_file_view(xl_path):
    pass


def render_settings_tab():
    st.write("Change Settings")


